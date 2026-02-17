#!/usr/bin/env python3
"""
Bank Financial Report CLI Tool
Tools untuk membuat laporan keuangan bank di Kali Linux Terminal
Integrasi langsung ke Excel dengan format profesional sesuai aturan akuntansi

Version: 1.0.0
"""

import argparse
import sys
import os
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field
from enum import Enum
import json

# Excel imports
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule


class TransactionType(Enum):
    """Jenis transaksi akuntansi"""
    SETORAN_TUNAI = "Setoran Tunai"
    PENARIKAN_TUNAI = "Penarikan Tunai"
    TRANSFER_MASUK = "Transfer Masuk"
    TRANSFER_KELUAR = "Transfer Keluar"
    PEMBAYARAN_TAGIHAN = "Pembayaran Tagihan"
    PEMBELIAN = "Pembelian/Investasi"
    BUNGA_MASUK = "Bunga Masuk"
    BIAYA_ADMIN = "Biaya Administrasi"
    PINJAMAN_MASUK = "Pinjaman Masuk"
    ANGSURAN_KELUAR = "Angsuran Keluar"


class AccountType(Enum):
    """Tipe akun akuntansi"""
    ASET = "Aset"
    KEWAJIBAN = "Kewajiban"
    EKUITAS = "Ekuitas"
    PENDAPATAN = "Pendapatan"
    BEBAN = "Beban"


@dataclass
class Transaction:
    """Data class untuk transaksi"""
    id: int
    date: str
    description: str
    transaction_type: TransactionType
    account_debit: str
    account_credit: str
    amount: float
    reference: str = ""
    notes: str = ""


@dataclass
class FinancialData:
    """Data class untuk menyimpan semua data keuangan"""
    bank_name: str = ""
    period: str = ""
    transactions: List[Transaction] = field(default_factory=list)
    
    # Saldo akun
    kas: float = 0.0
    bank_account: float = 0.0
    piutang: float = 0.0
    investasi: float = 0.0
    aset_lain: float = 0.0
    
    # Kewajiban
    hutang: float = 0.0
    pinjaman: float = 0.0
    
    # Ekuitas
    modal: float = 0.0
    laba_ditahan: float = 0.0
    
    # Pendapatan & Beban
    pendapatan_bunga: float = 0.0
    pendapatan_lain: float = 0.0
    beban_admin: float = 0.0
    beban_lain: float = 0.0


class BankFinancialReport:
    """Class utama untuk laporan keuangan bank"""
    
    def __init__(self):
        self.data = FinancialData()
        self.transaction_counter = 0
        self.setup_styles()
        
    def setup_styles(self):
        """Setup style untuk Excel"""
        # Colors - Professional Finance Style
        self.header_dark_blue = "1F4E79"
        self.header_light_blue = "2E75B6"
        self.accent_warm = "FFF3E0"
        self.bg_light = "F5F5F5"
        self.text_dark = "000000"
        self.positive_green = "27AE60"
        self.negative_red = "E74C3C"
        self.border_color = "B8C5D0"
        
        # Border style
        self.thin_border = Border(
            left=Side(style='thin', color=self.border_color),
            right=Side(style='thin', color=self.border_color),
            top=Side(style='thin', color=self.border_color),
            bottom=Side(style='thin', color=self.border_color)
        )
        
        self.header_fill = PatternFill(start_color=self.header_dark_blue, 
                                       end_color=self.header_dark_blue, fill_type="solid")
        self.subheader_fill = PatternFill(start_color=self.header_light_blue, 
                                          end_color=self.header_light_blue, fill_type="solid")
        self.alt_row_fill = PatternFill(start_color=self.bg_light, 
                                        end_color=self.bg_light, fill_type="solid")
        self.highlight_fill = PatternFill(start_color=self.accent_warm, 
                                          end_color=self.accent_warm, fill_type="solid")
        
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.title_font = Font(color=self.text_dark, bold=True, size=16)
        self.subtitle_font = Font(color=self.text_dark, bold=True, size=12)
        self.normal_font = Font(color=self.text_dark, size=10)
        self.currency_font = Font(color=self.text_dark, size=10)
        
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
    
    def show_banner(self):
        """Tampilkan banner aplikasi"""
        banner = """
╔══════════════════════════════════════════════════════════════════╗
║                                                                  ║
║           🏦 BANK FINANCIAL REPORT CLI TOOL 🏦                   ║
║                                                                  ║
║     Sistem Laporan Keuangan Bank dengan Integrasi Excel          ║
║              Sesuai Standar Akuntansi Indonesia                  ║
║                                                                  ║
╚══════════════════════════════════════════════════════════════════╝
        """
        print(banner)
    
    def show_menu(self):
        """Tampilkan menu utama"""
        menu = """
┌─────────────────────────────────────────────────────────────────┐
│                      MENU UTAMA                                 │
├─────────────────────────────────────────────────────────────────┤
│  [1] Input Transaksi Baru                                       │
│  [2] Lihat Daftar Transaksi                                     │
│  [3] Edit Transaksi                                             │
│  [4] Hapus Transaksi                                            │
│  [5] Set Saldo Awal                                             │
│  [6] Generate Laporan Keuangan (Excel)                          │
│  [7] Setting Informasi Bank                                     │
│  [0] Keluar                                                     │
└─────────────────────────────────────────────────────────────────┘
        """
        print(menu)
    
    def get_transaction_menu(self):
        """Menu pilihan jenis transaksi"""
        menu = """
┌─────────────────────────────────────────────────────────────────┐
│                  PILIH JENIS TRANSAKSI                          │
├─────────────────────────────────────────────────────────────────┤
│  [1] Setoran Tunai                    (Kas ↑ | Modal ↑)        │
│  [2] Penarikan Tunai                  (Kas ↓ | Modal ↓)        │
│  [3] Transfer Masuk                   (Bank ↑ | Pendapatan ↑)    │
│  [4] Transfer Keluar                  (Bank ↓ | Beban ↑)         │
│  [5] Pembayaran Tagihan               (Kas/Bank ↓ | Hutang ↓)    │
│  [6] Pembelian/Investasi              (Aset ↑ | Kas/Bank ↓)      │
│  [7] Bunga Masuk                      (Bank ↑ | Pendapatan ↑)    │
│  [8] Biaya Administrasi               (Beban ↑ | Bank ↓)         │
│  [9] Pinjaman Masuk                   (Bank ↑ | Pinjaman ↑)      │
│ [10] Angsuran Keluar                  (Pinjaman ↓ | Bank ↓)      │
│  [0] Kembali ke Menu Utama                                      │
└─────────────────────────────────────────────────────────────────┘
        """
        print(menu)
    
    def input_transaction(self):
        """Input transaksi baru"""
        self.get_transaction_menu()
        
        try:
            choice = input("Pilih jenis transaksi [0-10]: ").strip()
            
            if choice == '0':
                return
            
            if choice not in [str(i) for i in range(1, 11)]:
                print("❌ Pilihan tidak valid!")
                return
            
            print("\n" + "─" * 60)
            print("📋 FORM INPUT TRANSAKSI")
            print("─" * 60)
            
            date = input("Tanggal (YYYY-MM-DD) [default: hari ini]: ").strip()
            if not date:
                date = datetime.now().strftime("%Y-%m-%d")
            
            description = input("Keterangan/Deskripsi: ").strip()
            if not description:
                print("❌ Keterangan wajib diisi!")
                return
            
            amount_str = input("Nominal (Rp): ").strip().replace('.', '').replace(',', '')
            try:
                amount = float(amount_str)
                if amount <= 0:
                    print("❌ Nominal harus lebih dari 0!")
                    return
            except ValueError:
                print("❌ Nominal tidak valid!")
                return
            
            reference = input("No. Referensi/Bukti [optional]: ").strip()
            notes = input("Catatan tambahan [optional]: ").strip()
            
            # Proses berdasarkan jenis transaksi
            self.process_transaction(choice, date, description, amount, reference, notes)
            
        except KeyboardInterrupt:
            print("\n\n⚠️ Input dibatalkan.")
            return
    
    def process_transaction(self, choice: str, date: str, description: str, 
                           amount: float, reference: str, notes: str):
        """Proses transaksi berdasarkan jenisnya"""
        
        self.transaction_counter += 1
        trans_id = self.transaction_counter
        
        transaction_map = {
            '1': (TransactionType.SETORAN_TUNAI, "Kas", "Modal"),
            '2': (TransactionType.PENARIKAN_TUNAI, "Modal", "Kas"),
            '3': (TransactionType.TRANSFER_MASUK, "Bank", "Pendapatan Operasional"),
            '4': (TransactionType.TRANSFER_KELUAR, "Beban Operasional", "Bank"),
            '5': (TransactionType.PEMBAYARAN_TAGIHAN, "Hutang", "Bank"),
            '6': (TransactionType.PEMBELIAN, "Aset/Investasi", "Bank"),
            '7': (TransactionType.BUNGA_MASUK, "Bank", "Pendapatan Bunga"),
            '8': (TransactionType.BIAYA_ADMIN, "Beban Administrasi", "Bank"),
            '9': (TransactionType.PINJAMAN_MASUK, "Bank", "Pinjaman"),
            '10': (TransactionType.ANGSURAN_KELUAR, "Pinjaman", "Bank"),
        }
        
        trans_type, acc_debit, acc_credit = transaction_map[choice]
        
        # Update saldo berdasarkan jenis transaksi
        if choice == '1':  # Setoran Tunai
            self.data.kas += amount
            self.data.modal += amount
        elif choice == '2':  # Penarikan Tunai
            self.data.kas -= amount
            self.data.modal -= amount
        elif choice == '3':  # Transfer Masuk
            self.data.bank_account += amount
            self.data.pendapatan_lain += amount
        elif choice == '4':  # Transfer Keluar
            self.data.bank_account -= amount
            self.data.beban_lain += amount
        elif choice == '5':  # Pembayaran Tagihan
            self.data.hutang -= amount
            self.data.bank_account -= amount
        elif choice == '6':  # Pembelian/Investasi
            self.data.investasi += amount
            self.data.bank_account -= amount
        elif choice == '7':  # Bunga Masuk
            self.data.bank_account += amount
            self.data.pendapatan_bunga += amount
        elif choice == '8':  # Biaya Admin
            self.data.beban_admin += amount
            self.data.bank_account -= amount
        elif choice == '9':  # Pinjaman Masuk
            self.data.bank_account += amount
            self.data.pinjaman += amount
        elif choice == '10':  # Angsuran Keluar
            self.data.pinjaman -= amount
            self.data.bank_account -= amount
        
        # Buat objek transaksi
        transaction = Transaction(
            id=trans_id,
            date=date,
            description=description,
            transaction_type=trans_type,
            account_debit=acc_debit,
            account_credit=acc_credit,
            amount=amount,
            reference=reference,
            notes=notes
        )
        
        self.data.transactions.append(transaction)
        
        print("\n" + "=" * 60)
        print("✅ TRANSAKSI BERHASIL DISIMPAN")
        print("=" * 60)
        print(f"ID Transaksi    : {trans_id}")
        print(f"Tanggal         : {date}")
        print(f"Jenis           : {trans_type.value}")
        print(f"Keterangan      : {description}")
        print(f"Debit           : {acc_debit}")
        print(f"Kredit          : {acc_credit}")
        print(f"Nominal         : Rp {amount:,.2f}")
        print(f"No. Referensi   : {reference if reference else '-'}")
        print("=" * 60)
    
    def view_transactions(self):
        """Lihat daftar transaksi"""
        if not self.data.transactions:
            print("\n⚠️ Belum ada transaksi tercatat.")
            return
        
        print("\n" + "=" * 100)
        print("📋 DAFTAR TRANSAKSI")
        print("=" * 100)
        print(f"{'ID':<5} {'Tanggal':<12} {'Jenis Transaksi':<20} {'Keterangan':<25} {'Nominal':>15} {'Ref':<10}")
        print("-" * 100)
        
        for t in self.data.transactions:
            print(f"{t.id:<5} {t.date:<12} {t.transaction_type.value:<20} {t.description[:25]:<25} "
                  f"Rp {t.amount:>12,.0f} {t.reference[:10]:<10}")
        
        print("-" * 100)
        print(f"Total Transaksi: {len(self.data.transactions)}")
        print("=" * 100)
    
    def edit_transaction(self):
        """Edit transaksi yang sudah ada"""
        self.view_transactions()
        
        if not self.data.transactions:
            return
        
        try:
            trans_id = int(input("\nMasukkan ID transaksi yang akan diedit: "))
            transaction = next((t for t in self.data.transactions if t.id == trans_id), None)
            
            if not transaction:
                print("❌ Transaksi tidak ditemukan!")
                return
            
            print(f"\n📋 Edit Transaksi #{trans_id}")
            print("(Kosongkan jika tidak ingin mengubah)")
            
            new_desc = input(f"Keterangan [{transaction.description}]: ").strip()
            if new_desc:
                transaction.description = new_desc
            
            new_amount = input(f"Nominal [Rp {transaction.amount:,.0f}]: ").strip()
            if new_amount:
                transaction.amount = float(new_amount.replace('.', '').replace(',', ''))
            
            new_ref = input(f"Referensi [{transaction.reference}]: ").strip()
            if new_ref:
                transaction.reference = new_ref
            
            print("✅ Transaksi berhasil diupdate!")
            
        except ValueError:
            print("❌ Input tidak valid!")
    
    def delete_transaction(self):
        """Hapus transaksi"""
        self.view_transactions()
        
        if not self.data.transactions:
            return
        
        try:
            trans_id = int(input("\nMasukkan ID transaksi yang akan dihapus: "))
            transaction = next((t for t in self.data.transactions if t.id == trans_id), None)
            
            if not transaction:
                print("❌ Transaksi tidak ditemukan!")
                return
            
            confirm = input(f"Yakin hapus transaksi #{trans_id}? [y/N]: ").lower()
            if confirm == 'y':
                self.data.transactions.remove(transaction)
                print("✅ Transaksi berhasil dihapus!")
            else:
                print("❌ Penghapusan dibatalkan.")
                
        except ValueError:
            print("❌ Input tidak valid!")
    
    def set_initial_balance(self):
        """Set saldo awal"""
        print("\n" + "=" * 60)
        print("💰 SET SALDO AWAL")
        print("=" * 60)
        
        try:
            kas = input(f"Saldo Kas [Rp {self.data.kas:,.0f}]: ").strip()
            if kas:
                self.data.kas = float(kas.replace('.', '').replace(',', ''))
            
            bank = input(f"Saldo Bank [Rp {self.data.bank_account:,.0f}]: ").strip()
            if bank:
                self.data.bank_account = float(bank.replace('.', '').replace(',', ''))
            
            piutang = input(f"Piutang [Rp {self.data.piutang:,.0f}]: ").strip()
            if piutang:
                self.data.piutang = float(piutang.replace('.', '').replace(',', ''))
            
            hutang = input(f"Hutang [Rp {self.data.hutang:,.0f}]: ").strip()
            if hutang:
                self.data.hutang = float(hutang.replace('.', '').replace(',', ''))
            
            modal = input(f"Modal [Rp {self.data.modal:,.0f}]: ").strip()
            if modal:
                self.data.modal = float(modal.replace('.', '').replace(',', ''))
            
            print("\n✅ Saldo awal berhasil diupdate!")
            
        except ValueError:
            print("❌ Input tidak valid!")
    
    def set_bank_info(self):
        """Set informasi bank"""
        print("\n" + "=" * 60)
        print("🏦 SETTING INFORMASI BANK")
        print("=" * 60)
        
        bank_name = input(f"Nama Bank [{self.data.bank_name or 'Bank Anda'}]: ").strip()
        if bank_name:
            self.data.bank_name = bank_name
        
        period = input(f"Periode Laporan [{self.data.period or datetime.now().strftime('%B %Y')}]: ").strip()
        if period:
            self.data.period = period
        
        print("\n✅ Informasi bank berhasil diupdate!")
    
    def generate_excel_report(self, filename: str = None):
        """Generate laporan keuangan dalam format Excel"""
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            bank_name = self.data.bank_name.replace(" ", "_") if self.data.bank_name else "Bank"
            filename = f"Laporan_Keuangan_{bank_name}_{timestamp}.xlsx"
        
        print(f"\n📊 Generating Excel report: {filename}...")
        
        wb = Workbook()
        
        # Sheet 1: Cover
        self._create_cover_sheet(wb)
        
        # Sheet 2: Jurnal Transaksi
        self._create_journal_sheet(wb)
        
        # Sheet 3: Neraca
        self._create_neraca_sheet(wb)
        
        # Sheet 4: Laba Rugi
        self._create_laba_rugi_sheet(wb)
        
        # Sheet 5: Arus Kas
        self._create_arus_kas_sheet(wb)
        
        # Save workbook
        wb.save(filename)
        print(f"✅ Laporan berhasil disimpan: {filename}")
        
        return filename
    
    def _create_cover_sheet(self, wb: Workbook):
        """Create cover sheet"""
        ws = wb.active
        ws.title = "Cover"
        ws.sheet_view.showGridLines = False
        
        bank_name = self.data.bank_name or "BANK ANDA"
        period = self.data.period or datetime.now().strftime("%B %Y")
        
        # Title
        ws.merge_cells('B3:G3')
        ws['B3'] = "LAPORAN KEUANGAN"
        ws['B3'].font = Font(size=24, bold=True, color=self.header_dark_blue)
        ws['B3'].alignment = self.center_align
        ws.row_dimensions[3].height = 35
        
        ws.merge_cells('B4:G4')
        ws['B4'] = bank_name.upper()
        ws['B4'].font = Font(size=20, bold=True, color=self.header_light_blue)
        ws['B4'].alignment = self.center_align
        ws.row_dimensions[4].height = 30
        
        ws.merge_cells('B5:G5')
        ws['B5'] = f"Periode: {period}"
        ws['B5'].font = Font(size=14, color="666666")
        ws['B5'].alignment = self.center_align
        
        # Key Metrics
        ws['B8'] = "RINGKASAN KEUANGAN"
        ws['B8'].font = self.subtitle_font
        
        metrics = [
            ("Total Aset", self._get_total_aset()),
            ("Total Kewajiban", self._get_total_kewajiban()),
            ("Total Ekuitas", self._get_total_ekuitas()),
            ("Laba/Rugi Bersih", self._get_laba_rugi()),
            ("Saldo Kas", self.data.kas),
            ("Saldo Bank", self.data.bank_account),
        ]
        
        row = 10
        for label, value in metrics:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = self.normal_font
            ws[f'E{row}'] = value
            ws[f'E{row}'].number_format = 'Rp #,##0.00'
            ws[f'E{row}'].font = Font(bold=True, size=11)
            ws[f'E{row}'].alignment = self.right_align
            
            if value >= 0:
                ws[f'E{row}'].font = Font(bold=True, size=11, color=self.positive_green)
            else:
                ws[f'E{row}'].font = Font(bold=True, size=11, color=self.negative_red)
            row += 1
        
        # Sheet Index
        ws[f'B{row+2}'] = "DAFTAR ISI"
        ws[f'B{row+2}'].font = self.subtitle_font
        
        sheets = [
            ("Jurnal Transaksi", "Daftar semua transaksi yang tercatat"),
            ("Neraca", "Laporan posisi keuangan (Aset = Kewajiban + Ekuitas)"),
            ("Laba Rugi", "Laporan laba rugi periode berjalan"),
            ("Arus Kas", "Laporan arus kas masuk dan keluar"),
        ]
        
        idx_row = row + 4
        for sheet_name, desc in sheets:
            ws[f'B{idx_row}'] = sheet_name
            ws[f'B{idx_row}'].font = Font(bold=True, size=10)
            ws[f'D{idx_row}'] = desc
            ws[f'D{idx_row}'].font = self.normal_font
            idx_row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 3
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 3
        ws.column_dimensions['G'].width = 3
    
    def _create_journal_sheet(self, wb: Workbook):
        """Create jurnal transaksi sheet"""
        ws = wb.create_sheet("Jurnal Transaksi")
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('B2:I2')
        ws['B2'] = "JURNAL TRANSAKSI"
        ws['B2'].font = self.title_font
        ws['B2'].alignment = self.left_align
        ws.row_dimensions[2].height = 30
        
        # Headers
        headers = ['No', 'Tanggal', 'Jenis Transaksi', 'Keterangan', 
                   'Akun Debit', 'Akun Kredit', 'Nominal', 'Referensi']
        
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # Data
        row = 5
        for t in self.data.transactions:
            ws.cell(row=row, column=2, value=t.id).border = self.thin_border
            ws.cell(row=row, column=3, value=t.date).border = self.thin_border
            ws.cell(row=row, column=4, value=t.transaction_type.value).border = self.thin_border
            ws.cell(row=row, column=5, value=t.description).border = self.thin_border
            ws.cell(row=row, column=6, value=t.account_debit).border = self.thin_border
            ws.cell(row=row, column=7, value=t.account_credit).border = self.thin_border
            
            nominal_cell = ws.cell(row=row, column=8, value=t.amount)
            nominal_cell.number_format = 'Rp #,##0.00'
            nominal_cell.border = self.thin_border
            nominal_cell.alignment = self.right_align
            
            ws.cell(row=row, column=9, value=t.reference).border = self.thin_border
            
            # Alternating row color
            if row % 2 == 0:
                for col in range(2, 10):
                    ws.cell(row=row, column=col).fill = self.alt_row_fill
            
            row += 1
        
        # Total
        ws.merge_cells(f'B{row}:G{row}')
        ws[f'B{row}'] = "TOTAL"
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].alignment = self.right_align
        
        total_formula = f"=SUM(H5:H{row-1})"
        total_cell = ws[f'H{row}']
        total_cell.value = total_formula
        total_cell.number_format = 'Rp #,##0.00'
        total_cell.font = Font(bold=True)
        total_cell.fill = self.highlight_fill
        
        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 15
    
    def _create_neraca_sheet(self, wb: Workbook):
        """Create neraca sheet"""
        ws = wb.create_sheet("Neraca")
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('B2:E2')
        ws['B2'] = "NERACA (BALANCE SHEET)"
        ws['B2'].font = self.title_font
        ws.row_dimensions[2].height = 30
        
        period = self.data.period or datetime.now().strftime("%B %Y")
        ws.merge_cells('B3:E3')
        ws['B3'] = f"Periode: {period}"
        ws['B3'].font = Font(size=11, color="666666")
        
        # ASET
        ws['B5'] = "ASET"
        ws['B5'].font = self.subtitle_font
        ws['B5'].fill = self.subheader_fill
        ws.merge_cells('B5:D5')
        
        aset_items = [
            ("Kas", self.data.kas),
            ("Bank", self.data.bank_account),
            ("Piutang", self.data.piutang),
            ("Investasi", self.data.investasi),
            ("Aset Lainnya", self.data.aset_lain),
        ]
        
        row = 6
        for label, value in aset_items:
            ws[f'B{row}'] = f"  {label}"
            ws[f'B{row}'].font = self.normal_font
            ws[f'D{row}'] = value
            ws[f'D{row}'].number_format = 'Rp #,##0.00'
            ws[f'D{row}'].alignment = self.right_align
            row += 1
        
        # Total Aset
        ws[f'B{row}'] = "TOTAL ASET"
        ws[f'B{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'] = self._get_total_aset()
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'].fill = self.highlight_fill
        ws[f'D{row}'].alignment = self.right_align
        
        # KEWAJIBAN
        row += 2
        ws[f'B{row}'] = "KEWAJIBAN"
        ws[f'B{row}'].font = self.subtitle_font
        ws[f'B{row}'].fill = self.subheader_fill
        ws.merge_cells(f'B{row}:D{row}')
        
        kewajiban_items = [
            ("Hutang", self.data.hutang),
            ("Pinjaman", self.data.pinjaman),
        ]
        
        row += 1
        for label, value in kewajiban_items:
            ws[f'B{row}'] = f"  {label}"
            ws[f'B{row}'].font = self.normal_font
            ws[f'D{row}'] = value
            ws[f'D{row}'].number_format = 'Rp #,##0.00'
            ws[f'D{row}'].alignment = self.right_align
            row += 1
        
        ws[f'B{row}'] = "TOTAL KEWAJIBAN"
        ws[f'B{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'] = self._get_total_kewajiban()
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'].fill = self.highlight_fill
        ws[f'D{row}'].alignment = self.right_align
        
        # EKUITAS
        row += 2
        ws[f'B{row}'] = "EKUITAS"
        ws[f'B{row}'].font = self.subtitle_font
        ws[f'B{row}'].fill = self.subheader_fill
        ws.merge_cells(f'B{row}:D{row}')
        
        ekuitas_items = [
            ("Modal", self.data.modal),
            ("Laba Ditahan", self.data.laba_ditahan),
            ("Laba/Rugi Berjalan", self._get_laba_rugi()),
        ]
        
        row += 1
        for label, value in ekuitas_items:
            ws[f'B{row}'] = f"  {label}"
            ws[f'B{row}'].font = self.normal_font
            ws[f'D{row}'] = value
            ws[f'D{row}'].number_format = 'Rp #,##0.00'
            ws[f'D{row}'].alignment = self.right_align
            row += 1
        
        ws[f'B{row}'] = "TOTAL EKUITAS"
        ws[f'B{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'] = self._get_total_ekuitas()
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'].fill = self.highlight_fill
        ws[f'D{row}'].alignment = self.right_align
        
        # TOTAL KEWAJIBAN + EKUITAS
        row += 1
        ws[f'B{row}'] = "TOTAL KEWAJIBAN + EKUITAS"
        ws[f'B{row}'].font = Font(bold=True, size=12)
        ws[f'D{row}'] = self._get_total_kewajiban() + self._get_total_ekuitas()
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=12, color=self.header_dark_blue)
        ws[f'D{row}'].fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
        ws[f'D{row}'].alignment = self.right_align
        
        # Verification
        row += 2
        total_aset = self._get_total_aset()
        total_passiva = self._get_total_kewajiban() + self._get_total_ekuitas()
        
        if abs(total_aset - total_passiva) < 0.01:
            ws[f'B{row}'] = "✓ Neraca Seimbang (Aset = Kewajiban + Ekuitas)"
            ws[f'B{row}'].font = Font(bold=True, color=self.positive_green)
        else:
            ws[f'B{row}'] = f"✗ Neraca Tidak Seimbang (Selisih: Rp {abs(total_aset - total_passiva):,.2f})"
            ws[f'B{row}'].font = Font(bold=True, color=self.negative_red)
        
        ws.merge_cells(f'B{row}:E{row}')
        
        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 3
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 3
    
    def _create_laba_rugi_sheet(self, wb: Workbook):
        """Create laba rugi sheet"""
        ws = wb.create_sheet("Laba Rugi")
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('B2:E2')
        ws['B2'] = "LAPORAN LABA RUGI"
        ws['B2'].font = self.title_font
        ws.row_dimensions[2].height = 30
        
        period = self.data.period or datetime.now().strftime("%B %Y")
        ws.merge_cells('B3:E3')
        ws['B3'] = f"Periode: {period}"
        ws['B3'].font = Font(size=11, color="666666")
        
        # PENDAPATAN
        ws['B5'] = "PENDAPATAN"
        ws['B5'].font = self.subtitle_font
        ws['B5'].fill = self.subheader_fill
        ws.merge_cells('B5:D5')
        
        pendapatan_items = [
            ("Pendapatan Bunga", self.data.pendapatan_bunga),
            ("Pendapatan Lainnya", self.data.pendapatan_lain),
        ]
        
        row = 6
        for label, value in pendapatan_items:
            ws[f'B{row}'] = f"  {label}"
            ws[f'B{row}'].font = self.normal_font
            ws[f'D{row}'] = value
            ws[f'D{row}'].number_format = 'Rp #,##0.00'
            ws[f'D{row}'].alignment = self.right_align
            row += 1
        
        ws[f'B{row}'] = "TOTAL PENDAPATAN"
        ws[f'B{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'] = self._get_total_pendapatan()
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=11, color=self.positive_green)
        ws[f'D{row}'].fill = self.highlight_fill
        ws[f'D{row}'].alignment = self.right_align
        
        # BEBAN
        row += 2
        ws[f'B{row}'] = "BEBAN"
        ws[f'B{row}'].font = self.subtitle_font
        ws[f'B{row}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        ws.merge_cells(f'B{row}:D{row}')
        
        beban_items = [
            ("Beban Administrasi", self.data.beban_admin),
            ("Beban Lainnya", self.data.beban_lain),
        ]
        
        row += 1
        for label, value in beban_items:
            ws[f'B{row}'] = f"  {label}"
            ws[f'B{row}'].font = self.normal_font
            ws[f'D{row}'] = value
            ws[f'D{row}'].number_format = 'Rp #,##0.00'
            ws[f'D{row}'].alignment = self.right_align
            row += 1
        
        ws[f'B{row}'] = "TOTAL BEBAN"
        ws[f'B{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'] = self._get_total_beban()
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=11, color=self.negative_red)
        ws[f'D{row}'].fill = self.highlight_fill
        ws[f'D{row}'].alignment = self.right_align
        
        # LABA/RUGI BERSIH
        row += 2
        ws[f'B{row}'] = "LABA/RUGI BERSIH"
        ws[f'B{row}'].font = Font(bold=True, size=12)
        laba_rugi = self._get_laba_rugi()
        ws[f'D{row}'] = laba_rugi
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        
        if laba_rugi >= 0:
            ws[f'D{row}'].font = Font(bold=True, size=12, color=self.positive_green)
            ws[f'B{row+1}'] = "✓ Periode ini menghasilkan LABA"
        else:
            ws[f'D{row}'].font = Font(bold=True, size=12, color=self.negative_red)
            ws[f'B{row+1}'] = "✗ Periode ini mengalami RUGI"
        
        ws[f'D{row}'].fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        ws[f'D{row}'].alignment = self.right_align
        
        ws[f'B{row+1}'].font = Font(italic=True, size=10)
        ws.merge_cells(f'B{row+1}:D{row+1}')
        
        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 3
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 3
    
    def _create_arus_kas_sheet(self, wb: Workbook):
        """Create arus kas sheet"""
        ws = wb.create_sheet("Arus Kas")
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('B2:E2')
        ws['B2'] = "LAPORAN ARUS KAS"
        ws['B2'].font = self.title_font
        ws.row_dimensions[2].height = 30
        
        period = self.data.period or datetime.now().strftime("%B %Y")
        ws.merge_cells('B3:E3')
        ws['B3'] = f"Periode: {period}"
        ws['B3'].font = Font(size=11, color="666666")
        
        # ARUS KAS DARI AKTIVITAS OPERASI
        ws['B5'] = "ARUS KAS DARI AKTIVITAS OPERASI"
        ws['B5'].font = self.subtitle_font
        ws['B5'].fill = self.subheader_fill
        ws.merge_cells('B5:D5')
        
        operasi_items = [
            ("Penerimaan dari pelanggan", self.data.pendapatan_lain),
            ("Pembayaran beban operasional", -self.data.beban_lain),
            ("Pembayaran biaya administrasi", -self.data.beban_admin),
        ]
        
        row = 6
        for label, value in operasi_items:
            ws[f'B{row}'] = f"  {label}"
            ws[f'B{row}'].font = self.normal_font
            ws[f'D{row}'] = value
            ws[f'D{row}'].number_format = 'Rp #,##0.00'
            ws[f'D{row}'].alignment = self.right_align
            row += 1
        
        ws[f'B{row}'] = "Kas Bersih dari Operasi"
        ws[f'B{row}'].font = Font(bold=True, size=11)
        kas_operasi = sum(v for _, v in operasi_items)
        ws[f'D{row}'] = kas_operasi
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'].fill = self.highlight_fill
        ws[f'D{row}'].alignment = self.right_align
        
        # ARUS KAS DARI AKTIVITAS INVESTASI
        row += 2
        ws[f'B{row}'] = "ARUS KAS DARI AKTIVITAS INVESTASI"
        ws[f'B{row}'].font = self.subtitle_font
        ws[f'B{row}'].fill = self.subheader_fill
        ws.merge_cells(f'B{row}:D{row}')
        
        row += 1
        ws[f'B{row}'] = "  Pembelian investasi"
        ws[f'B{row}'].font = self.normal_font
        ws[f'D{row}'] = -self.data.investasi
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].alignment = self.right_align
        
        row += 1
        ws[f'B{row}'] = "  Penerimaan bunga"
        ws[f'B{row}'].font = self.normal_font
        ws[f'D{row}'] = self.data.pendapatan_bunga
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].alignment = self.right_align
        
        row += 1
        ws[f'B{row}'] = "Kas Bersih dari Investasi"
        ws[f'B{row}'].font = Font(bold=True, size=11)
        kas_investasi = self.data.pendapatan_bunga - self.data.investasi
        ws[f'D{row}'] = kas_investasi
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'].fill = self.highlight_fill
        ws[f'D{row}'].alignment = self.right_align
        
        # ARUS KAS DARI AKTIVITAS PENDANAAN
        row += 2
        ws[f'B{row}'] = "ARUS KAS DARI AKTIVITAS PENDANAAN"
        ws[f'B{row}'].font = self.subtitle_font
        ws[f'B{row}'].fill = self.subheader_fill
        ws.merge_cells(f'B{row}:D{row}')
        
        pendanaan_items = [
            ("  Setoran modal", self.data.modal),
            ("  Penerimaan pinjaman", self.data.pinjaman),
            ("  Pembayaran hutang", -self.data.hutang),
        ]
        
        row += 1
        for label, value in pendanaan_items:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = self.normal_font
            ws[f'D{row}'] = value
            ws[f'D{row}'].number_format = 'Rp #,##0.00'
            ws[f'D{row}'].alignment = self.right_align
            row += 1
        
        ws[f'B{row}'] = "Kas Bersih dari Pendanaan"
        ws[f'B{row}'].font = Font(bold=True, size=11)
        kas_pendanaan = sum(v for _, v in pendanaan_items)
        ws[f'D{row}'] = kas_pendanaan
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=11)
        ws[f'D{row}'].fill = self.highlight_fill
        ws[f'D{row}'].alignment = self.right_align
        
        # KENAIKAN/PENURUNAN KAS BERSIH
        row += 2
        ws[f'B{row}'] = "KENAIKAN/PENURUNAN KAS BERSIH"
        ws[f'B{row}'].font = Font(bold=True, size=12)
        total_arus_kas = kas_operasi + kas_investasi + kas_pendanaan
        ws[f'D{row}'] = total_arus_kas
        ws[f'D{row}'].number_format = 'Rp #,##0.00'
        ws[f'D{row}'].font = Font(bold=True, size=12, color=self.header_dark_blue)
        ws[f'D{row}'].fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
        ws[f'D{row}'].alignment = self.right_align
        
        # Set column widths
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 3
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 3
    
    def _get_total_aset(self) -> float:
        """Hitung total aset"""
        return self.data.kas + self.data.bank_account + self.data.piutang + self.data.investasi + self.data.aset_lain
    
    def _get_total_kewajiban(self) -> float:
        """Hitung total kewajiban"""
        return self.data.hutang + self.data.pinjaman
    
    def _get_total_ekuitas(self) -> float:
        """Hitung total ekuitas"""
        return self.data.modal + self.data.laba_ditahan + self._get_laba_rugi()
    
    def _get_total_pendapatan(self) -> float:
        """Hitung total pendapatan"""
        return self.data.pendapatan_bunga + self.data.pendapatan_lain
    
    def _get_total_beban(self) -> float:
        """Hitung total beban"""
        return self.data.beban_admin + self.data.beban_lain
    
    def _get_laba_rugi(self) -> float:
        """Hitung laba/rugi bersih"""
        return self._get_total_pendapatan() - self._get_total_beban()
    
    def run_interactive(self):
        """Run aplikasi dalam mode interaktif"""
        self.show_banner()
        
        # Set default bank info
        if not self.data.bank_name:
            self.data.bank_name = "Bank Saya"
        if not self.data.period:
            self.data.period = datetime.now().strftime("%B %Y")
        
        while True:
            self.show_menu()
            try:
                choice = input("Pilih menu [0-7]: ").strip()
                
                if choice == '0':
                    print("\n👋 Terima kasih telah menggunakan Bank Financial Report CLI!")
                    sys.exit(0)
                elif choice == '1':
                    self.input_transaction()
                elif choice == '2':
                    self.view_transactions()
                elif choice == '3':
                    self.edit_transaction()
                elif choice == '4':
                    self.delete_transaction()
                elif choice == '5':
                    self.set_initial_balance()
                elif choice == '6':
                    filename = input("Nama file output [default: auto]: ").strip()
                    if not filename:
                        filename = None
                    else:
                        if not filename.endswith('.xlsx'):
                            filename += '.xlsx'
                    self.generate_excel_report(filename)
                elif choice == '7':
                    self.set_bank_info()
                else:
                    print("❌ Pilihan tidak valid!")
                
                input("\nTekan Enter untuk melanjutkan...")
                
            except KeyboardInterrupt:
                print("\n\n👋 Terima kasih telah menggunakan Bank Financial Report CLI!")
                sys.exit(0)


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Bank Financial Report CLI Tool - Laporan Keuangan Bank dengan Integrasi Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Contoh Penggunaan:
  %(prog)s                          # Jalankan mode interaktif
  %(prog)s --output laporan.xlsx    # Generate laporan dengan nama file custom
  %(prog)s --demo                   # Jalankan dengan data demo

Author: Financial CLI Tools
Version: 1.0.0
        """
    )
    
    parser.add_argument('-o', '--output', 
                       help='Nama file output Excel (default: auto-generated)')
    parser.add_argument('--demo', action='store_true',
                       help='Jalankan dengan data demo untuk testing')
    parser.add_argument('-v', '--version', action='version', version='%(prog)s 1.0.0')
    
    args = parser.parse_args()
    
    app = BankFinancialReport()
    
    if args.demo:
        # Load demo data
        app.data.bank_name = "Bank Demo Indonesia"
        app.data.period = datetime.now().strftime("%B %Y")
        app.data.kas = 50000000
        app.data.bank_account = 250000000
        app.data.piutang = 75000000
        app.data.investasi = 100000000
        app.data.modal = 400000000
        app.data.pinjaman = 75000000
        
        # Add demo transactions
        demo_trans = [
            ("2024-01-05", TransactionType.SETORAN_TUNAI, "Setoran Modal Awal", "Kas", "Modal", 100000000),
            ("2024-01-10", TransactionType.BUNGA_MASUK, "Bunga Deposito", "Bank", "Pendapatan Bunga", 2500000),
            ("2024-01-15", TransactionType.PEMBELIAN, "Pembelian Obligasi", "Investasi", "Bank", 50000000),
            ("2024-01-20", TransactionType.BIAYA_ADMIN, "Biaya Admin Bulanan", "Beban Administrasi", "Bank", 150000),
            ("2024-01-25", TransactionType.TRANSFER_MASUK, "Penerimaan Transfer", "Bank", "Pendapatan", 15000000),
        ]
        
        for i, (date, t_type, desc, acc_d, acc_c, amount) in enumerate(demo_trans, 1):
            app.transaction_counter += 1
            app.data.transactions.append(Transaction(
                id=app.transaction_counter,
                date=date,
                description=desc,
                transaction_type=t_type,
                account_debit=acc_d,
                account_credit=acc_c,
                amount=amount
            ))
        
        print("✅ Data demo berhasil dimuat!")
        app.run_interactive()
    elif args.output:
        app.generate_excel_report(args.output)
    else:
        app.run_interactive()


if __name__ == "__main__":
    main()
